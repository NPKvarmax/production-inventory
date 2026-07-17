from django.db.models import Sum, Count
from django.db import transaction 
import json
import csv
import openpyxl
from datetime import datetime, date
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from .models import InventoryItem, StockRequest, BuildKit
from django.contrib.auth.forms import AuthenticationForm
from django.contrib.auth import login, logout

from django.http import HttpResponse
from django.utils.timezone import now
from .forms import StockRequestForm

@login_required(login_url='login')
def landing_page(request):
    context = {
        'username': request.user.username
    }
    return render(request, 'inventory/landing.html', context)

@login_required(login_url='login') 
def dashboard(request, item_type):
    
    if item_type not in ['Part', 'Fastener']:
        return redirect('landing')

    items = InventoryItem.objects.filter(item_type=item_type)
    kits = BuildKit.objects.filter(components__item__item_type=item_type).distinct()
    categories = items.values_list('category', flat=True).distinct()
    
    category_filter = request.GET.get('category')
    reorder_filter = request.GET.get('reorder')
    priority_filter = request.GET.get('priority')
    station_filter = request.GET.get('station')
    sort_filter = request.GET.get('sort')
    search_query = request.GET.get('search', '').strip() 
    
    if search_query:
        items = items.filter(name__icontains=search_query)
    if category_filter and category_filter != 'All':
        items = items.filter(category=category_filter)
    if station_filter and station_filter != 'All':
        items = items.filter(station=station_filter)
    if priority_filter and priority_filter != 'All':
        if not isinstance(items, list):
            items = items.filter(stockrequest__priority=priority_filter).distinct()
    if reorder_filter == 'Yes':
        items = [item for item in items if item.reorder_needed == 'Yes']
    elif reorder_filter == 'No':
        items = [item for item in items if item.reorder_needed == 'No']
        
    if not isinstance(items, list):
        items = list(items)
        
    if sort_filter == 'value_desc':
        items.sort(key=lambda x: x.total_value if hasattr(x, 'total_value') else 0, reverse=True)
    
    if request.user.is_superuser:
        recent_requests = StockRequest.objects.filter(item__item_type=item_type).order_by('-date_requested')
    else:
        recent_requests = StockRequest.objects.filter(item__item_type=item_type, requester=request.user).order_by('-date_requested')    

    if priority_filter and priority_filter != 'All':
        recent_requests = recent_requests.filter(priority=priority_filter)
    
    recent_requests = recent_requests[:10]
    
    total_items_count = len(items) if isinstance(items, list) else items.count()
    low_stock_count = sum(1 for item in items if item.reorder_needed == 'Yes')
    total_inventory_value = sum(item.total_value for item in items)
    defect_stats = StockRequest.objects.filter(
        item__item_type=item_type, 
        priority='Defect Replacement'
    ).aggregate(total=Sum('quantity_requested'))
    defect_count = defect_stats['total'] or 0

    chart_labels = json.dumps([item.name for item in items])
    chart_data = json.dumps([float(item.quantity) for item in items]) 

    if request.method == 'POST':
        
        #  LIVE AUDIT (PHYSICAL COUNT) OPERATION
        if 'audit_submit' in request.POST:
            item_id = request.POST.get('audit_item_id')
            physical_qty = int(request.POST.get('audit_qty', 0))
            audit_item = get_object_or_404(InventoryItem, id=item_id)
            system_qty = audit_item.quantity
            variance = physical_qty - system_qty

            if variance == 0:
                priority_type = 'Audit Match'
                logged_qty = 0
            elif variance < 0:
                priority_type = 'Audit Shortage'
                logged_qty = abs(variance)
            else:
                priority_type = 'Audit Surplus'
                logged_qty = variance

            audit_item.quantity = physical_qty
            audit_item.save()

            StockRequest.objects.create(
                item=audit_item, quantity_requested=logged_qty,
                requester=request.user, priority=priority_type
            )
            
            if priority_type == 'Audit Match':
                messages.success(request, f"Perfect Match! {audit_item.name} verified at {physical_qty} units.")
            else:
                messages.warning(request, f"Discrepancy logged. {audit_item.name} updated from {system_qty} to {physical_qty}.")
            return redirect('dashboard', item_type=item_type)

        #  RESTOCK OPERATION
        elif 'restock_submit' in request.POST:
            item_id = request.POST.get('restock_item_id')
            qty_to_add = int(request.POST.get('restock_qty', 0))
            restock_item = get_object_or_404(InventoryItem, id=item_id)
            restock_item.quantity += qty_to_add
            restock_item.save()

            StockRequest.objects.create(
                item=restock_item, quantity_requested=qty_to_add,
                requester=request.user, priority='Restock (Incoming)' 
            )
            messages.success(request, f"Successfully restocked {qty_to_add} units of {restock_item.name}.")
            return redirect('dashboard', item_type=item_type)

        #  KIT REQUESTS 
        elif 'request_kit' in request.POST:
            kit_id = request.POST.get('kit_id')
            qty_to_build = int(request.POST.get('kit_quantity', 1))
            kit = get_object_or_404(BuildKit, id=kit_id)

            can_build = True
            missing_parts = []
            
            for component in kit.components.all():
                total_needed = component.quantity_required * qty_to_build
                if component.item.quantity < total_needed:
                    can_build = False
                    missing_parts.append(f"{component.item.name} (Need {total_needed}, Have {component.item.quantity})")

            if can_build:
                for component in kit.components.all():
                    total_needed = component.quantity_required * qty_to_build
                    component.item.quantity -= total_needed
                    component.item.save()
                    
                    StockRequest.objects.create(
                        item=component.item, quantity_requested=total_needed,
                        requester=request.user
                    )
                messages.success(request, f"Successfully pulled parts for {qty_to_build}x {kit.name}!")
            else:
                messages.error(request, f"Cannot build {kit.name}. Missing parts: {', '.join(missing_parts)}")
            return redirect('dashboard', item_type=item_type)

        # 4. BULK CATALOGUE REQUESTS (NEW!)
        elif 'bulk_request_submit' in request.POST:
            cart_data_raw = request.POST.get('cart_data', '[]')
            try:
                cart_data = json.loads(cart_data_raw)
                if not cart_data:
                    messages.error(request, "Your request list is empty.")
                    return redirect('dashboard', item_type=item_type)
                
                # transaction.atomic() so if ONE item fails, the WHOLE cart is cancelled.
                with transaction.atomic():
                    # Pre-checking all items to make sure we have enough stock
                    for item_data in cart_data:
                        db_item = InventoryItem.objects.select_for_update().get(id=item_data['id'])
                        qty = int(item_data['qty'])
                        if qty > db_item.quantity:
                            raise ValueError(f"Not enough stock for {db_item.name}. You asked for {qty}, but only {db_item.quantity} are left.")
                    
                    # After check, deduct stock and create logs
                    for item_data in cart_data:
                        db_item = InventoryItem.objects.get(id=item_data['id'])
                        qty = int(item_data['qty'])
                        
                        db_item.quantity -= qty
                        db_item.save()
                        
                        StockRequest.objects.create(
                            item=db_item,
                            quantity_requested=qty,
                            priority=item_data['priority'],
                            requester=request.user
                        )
                messages.success(request, f"Successfully requested {len(cart_data)} items from the catalogue.")
            except ValueError as e:
                # "Not enough stock" error
                messages.error(request, str(e))
            except Exception as e:
                messages.error(request, "An unexpected error occurred while processing your list.")
                
            return redirect('dashboard', item_type=item_type)

    available_items = InventoryItem.objects.filter(item_type=item_type)
    
    context = {
        'item_type': item_type, 
        'items': items,
        'kits': kits, 
        'available_items': available_items,
        'recent_requests': recent_requests,
        'categories': categories,
        'total_items_count': total_items_count,
        'low_stock_count': low_stock_count,
        'total_inventory_value': total_inventory_value,
        'defect_count': defect_count,
        'chart_labels': chart_labels,
        'chart_data': chart_data,
        'search_query': search_query,
    }
    return render(request, 'inventory/dashboard.html', context)


def login_page(request):
    if request.user.is_authenticated:
        return redirect('landing')

    if request.method == 'POST':
        form = AuthenticationForm(request, data=request.POST)
        if form.is_valid():
            user = form.get_user()
            login(request, user)
            return redirect('landing') 
    else:
        form = AuthenticationForm()

    context = {'form': form}
    return render(request, 'inventory/login.html', context)


def logout_user(request):
    logout(request)
    return redirect('login')


@login_required(login_url='login')
def export_inventory_csv(request):
    if not request.user.is_superuser:
        messages.error(request, "Access Denied: Only administrators can download reports.")
        return redirect('landing')

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="complete_inventory_status.csv"'

    writer = csv.writer(response)
    writer.writerow(['Type', 'Item Name', 'Current Quantity', 'Status'])

    items = InventoryItem.objects.all()
    for item in items:
        status = 'LOW STOCK' if item.quantity < 10 else 'OK'
        writer.writerow([item.item_type, item.name, item.quantity, status])

    return response


# EXECUTIVE SUMMARY REPORT
@login_required(login_url='login')
def executive_report(request):
    if not request.user.is_superuser:
        messages.error(request, "Access Denied: Executive reporting only.")
        return redirect('landing')

    report_period = request.GET.get('report_period')
    current_date = now()
    if report_period:
        try:
            target_year, target_month = map(int, report_period.split('-'))
        except ValueError:
            target_year, target_month = current_date.year, current_date.month
    else:
        target_year, target_month = current_date.year, current_date.month

    report_month_name = date(target_year, target_month, 1).strftime("%B %Y")

    all_items = InventoryItem.objects.all()
    total_units = sum(item.quantity for item in all_items)
    total_value = sum(item.total_value for item in all_items if hasattr(item, 'total_value')) 

    is_historical = (target_year < current_date.year) or (target_year == current_date.year and target_month < current_date.month)

    if is_historical:
        if target_month == 12:
            cutoff_date = datetime(target_year + 1, 1, 1)
        else:
            cutoff_date = datetime(target_year, target_month + 1, 1)
            
        future_requests = StockRequest.objects.filter(date_requested__gte=cutoff_date).select_related('item')
        
        for req in future_requests:
            if req.priority in ['Restock (Incoming)', 'Audit Surplus']:
                total_units -= req.quantity_requested
                total_value -= (req.quantity_requested * req.item.unit_cost)
            elif req.priority == 'Audit Match':
                pass 
            else:
                total_units += req.quantity_requested
                total_value += (req.quantity_requested * req.item.unit_cost)

    out_of_stock = InventoryItem.objects.filter(quantity__lte=0)
    low_stock = InventoryItem.objects.filter(quantity__lte=10)
    
    monthly_requests = StockRequest.objects.filter(
        date_requested__year=target_year,
        date_requested__month=target_month
    )
    
    consumption_requests = monthly_requests.exclude(
        priority__in=['Restock (Incoming)', 'Audit Match', 'Audit Shortage', 'Audit Surplus']
    )

    fast_moving_items = consumption_requests.values('item__name').annotate(
        total_pulled=Sum('quantity_requested')
    ).order_by('-total_pulled')[:20]

    defect_replacements = consumption_requests.filter(priority='Defect Replacement').values('item__name').annotate(
        total_replaced=Sum('quantity_requested')
    ).order_by('-total_replaced')

    priority_breakdown = consumption_requests.values('priority').annotate(
        request_count=Count('id'),
        total_volume=Sum('quantity_requested')
    )

    restock_logs = monthly_requests.filter(priority='Restock (Incoming)').values('item__name').annotate(
        total_added=Sum('quantity_requested')
    ).order_by('-total_added')

    audit_logs = monthly_requests.filter(priority__startswith='Audit')
    total_audits = audit_logs.count()
    total_matches = audit_logs.filter(priority='Audit Match').count()
    
    accuracy_rate = None
    if total_audits > 0:
        accuracy_rate = (total_matches / total_audits) * 100

    audit_discrepancies = audit_logs.exclude(priority='Audit Match').values(
        'item__name', 'priority'
    ).annotate(
        variance=Sum('quantity_requested')
    ).order_by('-variance')

    live_deficits = []
    total_unmatched_deficit = 0

    frame = InventoryItem.objects.filter(name__icontains='FRAME').first()
    if frame:
        frame_children_names = [
            'LEFT CRANK', 'RIGHT CRANK', 'CHAINRING', 'REAR DERAILLEUR', 
            'TORQUE SENSOR', 'KICKSTAND', 'REAR MUDGUARD', 'FENDER STAY', 'REAR SHOCK'
        ]
        for part in InventoryItem.objects.filter(name__in=frame_children_names):
            if part.quantity < frame.quantity:
                gap = frame.quantity - part.quantity
                live_deficits.append({'part': part.name, 'gap': gap, 'parent': frame.name})
                total_unmatched_deficit += gap

    handlebar = InventoryItem.objects.filter(name__icontains='HANDLEBAR').first()
    if handlebar:
        handlebar_children_names = [
            'GRIPS', 'THROTTLE', 'DISPLAY (MONITOR)', 'FRONT BRAKE', 
            'REAR BRAKE', 'HANDLEBAR SWITCH', 'GEAR SHIFTER', 'MIRROR'
        ]
        for part in InventoryItem.objects.filter(name__in=handlebar_children_names):
            if part.quantity < handlebar.quantity:
                gap = handlebar.quantity - part.quantity
                live_deficits.append({'part': part.name, 'gap': gap, 'parent': handlebar.name})
                total_unmatched_deficit += gap

    front_hub = InventoryItem.objects.filter(name__icontains='FRONT HUB').first()
    if front_hub:
        for part in InventoryItem.objects.filter(name__in=['FRONT TUBE', 'FRONT TYRE']):
            if part.quantity < front_hub.quantity:
                gap = front_hub.quantity - part.quantity
                live_deficits.append({'part': part.name, 'gap': gap, 'parent': front_hub.name})
                total_unmatched_deficit += gap
        
        front_spokes = InventoryItem.objects.filter(name__icontains='FRONT SPOKE').first()
        if front_spokes:
            expected_spokes = front_hub.quantity * 36
            if front_spokes.quantity < expected_spokes:
                gap = expected_spokes - front_spokes.quantity
                live_deficits.append({'part': front_spokes.name, 'gap': gap, 'parent': f"{front_hub.name} (Requires 36x)"})
                total_unmatched_deficit += gap

    motor = InventoryItem.objects.filter(name__icontains='MOTOR').exclude(name__icontains='CABLE').first()
    if motor:
        for part in InventoryItem.objects.filter(name__in=['REAR TUBE', 'REAR TYRE']):
            if part.quantity < motor.quantity:
                gap = motor.quantity - part.quantity
                live_deficits.append({'part': part.name, 'gap': gap, 'parent': motor.name})
                total_unmatched_deficit += gap
        
        rear_spokes = InventoryItem.objects.filter(name__icontains='REAR SPOKE').first()
        if rear_spokes:
            expected_spokes = motor.quantity * 36
            if rear_spokes.quantity < expected_spokes:
                gap = expected_spokes - rear_spokes.quantity
                live_deficits.append({'part': rear_spokes.name, 'gap': gap, 'parent': f"{motor.name} (Requires 36x)"})
                total_unmatched_deficit += gap

    balancer = InventoryItem.objects.filter(name__icontains='BALANCER MODULE').first()
    if balancer:
        balancer_children = [
            'CHAIN', 'PEDALS', 'STEM', 'CABLE SET: DERAILLEUR OUTER CABLE', 
            'BATTERY MOUNT C', 'BATTERY MOUNT D', 'CONTROLLER', 
            'EB-BUS CABLE 1', 'EB-BUS CABLE 2', 'MOTOR EXTENSION CABLE'
        ]
        for part in InventoryItem.objects.filter(name__in=balancer_children):
            if part.quantity < balancer.quantity:
                gap = balancer.quantity - part.quantity
                live_deficits.append({'part': part.name, 'gap': gap, 'parent': balancer.name})
                total_unmatched_deficit += gap

    fork = InventoryItem.objects.filter(name__icontains='FORK').first()
    if fork:
        fork_children = [
            'FRONT LIGHT', 'HORN', 'FRONT MUDGUARD', 
            'FRONT LIGHT BRACKET SET', 'TORQUE SENSOR CABLE'
        ]
        for part in InventoryItem.objects.filter(name__in=fork_children):
            if part.quantity < fork.quantity:
                gap = fork.quantity - part.quantity
                live_deficits.append({'part': part.name, 'gap': gap, 'parent': fork.name})
                total_unmatched_deficit += gap

    context = {
        'total_value': total_value,
        'total_units': total_units,
        'out_of_stock': out_of_stock,
        'low_stock': low_stock,
        'fast_moving_items': fast_moving_items,
        'defect_replacements': defect_replacements,
        'priority_breakdown': priority_breakdown,
        'restock_logs': restock_logs,
        'current_month_name': report_month_name,
        'selected_period': f"{target_year}-{target_month:02d}",
        'total_requests': monthly_requests.count(),
        'live_deficits': live_deficits,
        'is_historical': is_historical,
        'accuracy_rate': accuracy_rate,
        'total_audits': total_audits,
        'audit_discrepancies': audit_discrepancies,
    }
    return render(request, 'inventory/executive_report.html', context)


@login_required(login_url='login')
def export_excel_report(request):
    if not request.user.is_superuser:
        messages.error(request, "Access Denied.")
        return redirect('landing')

    report_period = request.GET.get('report_period')
    if report_period:
        try:
            target_year, target_month = map(int, report_period.split('-'))
        except ValueError:
            target_year, target_month = now().year, now().month
    else:
        target_year, target_month = now().year, now().month

    report_month_name = date(target_year, target_month, 1).strftime("%b_%Y")

    wb = openpyxl.Workbook()
    
    ws1 = wb.active
    ws1.title = "Inventory Overview"
    ws1.append(["Company", "Wahu Mobility"])
    ws1.append(["Report Date", datetime.now().strftime("%Y-%m-%d")])
    ws1.append([]) 
    
    ws1.append(["Item Name", "Category", "Station", "Quantity", "Total Value"])
    items = InventoryItem.objects.all()
    for item in items:
        val = item.total_value if hasattr(item, 'total_value') else 0 
        ws1.append([item.name, item.category, item.station, item.quantity, val])

    ws2 = wb.create_sheet(title=f"Logs {report_month_name}")
    ws2.append(["Date Requested", "Item Name", "Priority/Type", "Quantity", "Requester"])
    
    monthly_requests = StockRequest.objects.filter(
        date_requested__year=target_year, 
        date_requested__month=target_month
    ).order_by('-date_requested')

    for req in monthly_requests:
        user_name = req.requester.username if req.requester else "System / Admin"
        ws2.append([
            req.date_requested.strftime("%Y-%m-%d"), 
            req.item.name, 
            req.priority,
            req.quantity_requested, 
            user_name
        ])

    ws3 = wb.create_sheet(title="Urgent Actions")
    ws3.append(["URGENT RESTOCK REQUIRED"])
    ws3.append(["Item Name", "Station", "Last Known Quantity"])
    
    critical_items = InventoryItem.objects.filter(quantity__lte=0)
    for item in critical_items:
        ws3.append([item.name, item.station, item.quantity])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="Wahu_Inventory_Report_{report_month_name}.xlsx"'
    
    wb.save(response)
    return response

# BOM RECONCILIATION & MRP DASHBOARD

@login_required(login_url='login')
def bom_reconciliation(request):
    if not request.user.is_superuser:
        messages.error(request, "Access Denied: Manager clearance required.")
        return redirect('landing')

    live_deficits = []
    total_unmatched_deficit = 0

    frame = InventoryItem.objects.filter(name__icontains='FRAME').first()
    if frame:
        frame_children_names = [
            'LEFT CRANK', 'RIGHT CRANK', 'CHAINRING', 'REAR DERAILLEUR', 
            'TORQUE SENSOR', 'KICKSTAND', 'REAR MUDGUARD', 'FENDER STAY', 'REAR SHOCK'
        ]
        for part in InventoryItem.objects.filter(name__in=frame_children_names):
            if part.quantity < frame.quantity:
                gap = frame.quantity - part.quantity
                live_deficits.append({'part': part.name, 'gap': gap, 'parent': frame.name, 'ratio': '1:1'})
                total_unmatched_deficit += gap

    handlebar = InventoryItem.objects.filter(name__icontains='HANDLEBAR').first()
    if handlebar:
        handlebar_children_names = [
            'GRIPS', 'THROTTLE', 'DISPLAY (MONITOR)', 'FRONT BRAKE', 
            'REAR BRAKE', 'HANDLEBAR SWITCH', 'GEAR SHIFTER', 'MIRROR'
        ]
        for part in InventoryItem.objects.filter(name__in=handlebar_children_names):
            if part.quantity < handlebar.quantity:
                gap = handlebar.quantity - part.quantity
                live_deficits.append({'part': part.name, 'gap': gap, 'parent': handlebar.name, 'ratio': '1:1'})
                total_unmatched_deficit += gap

    front_hub = InventoryItem.objects.filter(name__icontains='FRONT HUB').first()
    if front_hub:
        for part in InventoryItem.objects.filter(name__in=['FRONT TUBE', 'FRONT TYRE']):
            if part.quantity < front_hub.quantity:
                gap = front_hub.quantity - part.quantity
                live_deficits.append({'part': part.name, 'gap': gap, 'parent': front_hub.name, 'ratio': '1:1'})
                total_unmatched_deficit += gap
        
        front_spokes = InventoryItem.objects.filter(name__icontains='FRONT SPOKE').first()
        if front_spokes:
            expected_spokes = front_hub.quantity * 36
            if front_spokes.quantity < expected_spokes:
                gap = expected_spokes - front_spokes.quantity
                live_deficits.append({'part': front_spokes.name, 'gap': gap, 'parent': front_hub.name, 'ratio': '36:1'})
                total_unmatched_deficit += gap

    motor = InventoryItem.objects.filter(name__icontains='MOTOR').exclude(name__icontains='CABLE').first()
    if motor:
        for part in InventoryItem.objects.filter(name__in=['REAR TUBE', 'REAR TYRE']):
            if part.quantity < motor.quantity:
                gap = motor.quantity - part.quantity
                live_deficits.append({'part': part.name, 'gap': gap, 'parent': motor.name, 'ratio': '1:1'})
                total_unmatched_deficit += gap
        
        rear_spokes = InventoryItem.objects.filter(name__icontains='REAR SPOKE').first()
        if rear_spokes:
            expected_spokes = motor.quantity * 36
            if rear_spokes.quantity < expected_spokes:
                gap = expected_spokes - rear_spokes.quantity
                live_deficits.append({'part': rear_spokes.name, 'gap': gap, 'parent': motor.name, 'ratio': '36:1'})
                total_unmatched_deficit += gap

    balancer = InventoryItem.objects.filter(name__icontains='BALANCER MODULE').first()
    if balancer:
        balancer_children = [
            'CHAIN', 'PEDALS', 'STEM', 'CABLE SET: DERAILLEUR OUTER CABLE', 
            'BATTERY MOUNT C', 'BATTERY MOUNT D', 'CONTROLLER', 
            'EB-BUS CABLE 1', 'EB-BUS CABLE 2', 'MOTOR EXTENSION CABLE'
        ]
        for part in InventoryItem.objects.filter(name__in=balancer_children):
            if part.quantity < balancer.quantity:
                gap = balancer.quantity - part.quantity
                live_deficits.append({'part': part.name, 'gap': gap, 'parent': balancer.name, 'ratio': '1:1'})
                total_unmatched_deficit += gap

    fork = InventoryItem.objects.filter(name__icontains='FORK').first()
    if fork:
        fork_children = [
            'FRONT LIGHT', 'HORN', 'FRONT MUDGUARD', 
            'FRONT LIGHT BRACKET SET', 'TORQUE SENSOR CABLE'
        ]
        for part in InventoryItem.objects.filter(name__in=fork_children):
            if part.quantity < fork.quantity:
                gap = fork.quantity - part.quantity
                live_deficits.append({'part': part.name, 'gap': gap, 'parent': fork.name, 'ratio': '1:1'})
                total_unmatched_deficit += gap

    context = {
        'live_deficits': live_deficits,
        'total_unmatched': total_unmatched_deficit
    }
    return render(request, 'inventory/reconciliation.html', context)