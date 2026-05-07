import openpyxl
import io
from datetime import datetime
from django.core.management.base import BaseCommand
from django.core.mail import EmailMessage
from django.conf import settings
from inventory.models import InventoryItem, StockRequest

class Command(BaseCommand):
    help = 'Generates the end-of-month inventory report and emails it to management.'

    def handle(self, *args, **kwargs):
        self.stdout.write("Starting to generate monthly report...")

        wb = openpyxl.Workbook()
        
        # --- OVERVIEW ---
        ws1 = wb.active
        ws1.title = "Inventory Overview"
        ws1.append(["Company", "Wahu Mobility"])
        ws1.append(["Report Date", datetime.now().strftime("%Y-%m-%d")])
        ws1.append([])
        ws1.append(["Item Name", "Category", "Station", "Quantity", "Total Value"])
        
        for item in InventoryItem.objects.all():
            val = item.total_value if hasattr(item, 'total_value') else 0 
            ws1.append([item.name, item.category, item.station, item.quantity, val])

        # --- MONTHLY CONSUMPTION ---
        ws2 = wb.create_sheet(title="Monthly Consumption")
        ws2.append(["Date Requested", "Item Name", "Quantity Pulled", "Requester"])
        
        current_month = datetime.now().month
        current_year = datetime.now().year
        for req in StockRequest.objects.filter(date_requested__year=current_year, date_requested__month=current_month):
            ws2.append([req.date_requested.strftime("%Y-%m-%d"), req.item.name, req.quantity_requested, req.requester.username])

        # --- URGENT ACTIONS ---
        ws3 = wb.create_sheet(title="Urgent Actions")
        ws3.append(["URGENT RESTOCK REQUIRED"])
        ws3.append(["Item Name", "Station", "Last Known Quantity"])
        
        for item in InventoryItem.objects.filter(quantity__lte=0):
            ws3.append([item.name, item.station, item.quantity])

        #  Saving the Excel file to the server's RAM (Memory)
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        filename = f'Wahu_Inventory_Report_{datetime.now().strftime("%b_%Y")}.xlsx'

        #  Creating and Sending the Email
        self.stdout.write("Excel file generated. Preparing email...")
        
        email = EmailMessage(
            subject=f'Automated Monthly Inventory Report - {datetime.now().strftime("%B %Y")}',
            body='Hello Management,\n\nPlease find the automated end-of-month provisional inventory report attached.\n\nBest,\nYour Automated Inventory Management System',
            from_email=settings.EMAIL_HOST_USER,
            to=['vopoku@wahu.me','abena@wahu.me','quincy@wahu.me'], 
        )
        
        # Attaching the file from memory
        email.attach(filename, excel_file.read(), 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        
        # Sending the email
        email.send()
        
        self.stdout.write(self.style.SUCCESS('Successfully sent the monthly report!'))